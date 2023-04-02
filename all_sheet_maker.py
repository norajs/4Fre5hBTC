# -*- coding: utf-8 -*-
"""
Created on Sun Apr  2 15:29:25 2023

@author: noraj
"""

import openpyxl
import os
import random

### Excel file for final lists
excel = "List.xlsx"
wb_obj = openpyxl.load_workbook(excel)
#Sheets
sheet_v1 = wb_obj["v1"]
sheet_v2 = wb_obj["v2"]
sheet_v3 = wb_obj["v3"]
sheet_v4 = wb_obj["v4"]
sheet_v5 = wb_obj["v5"]
sheet_all = wb_obj["all"]

row_counter = 2
col_counter = 1

#for file in dirs:
#current_sheet['A' + str(row_count)] = htmlName
#current_sheet.cell(row=row_count, column=col_count).value = d_temp



for row in range(2, 85):
    cell = sheet_v1.cell(row=row_counter, column=1)
    if cell.value:
        
    