# -*- coding: utf-8 -*-
"""
Created on Fri Mar 24 10:10:56 2023

@author: noraj
"""

import openpyxl
import os
import random

### Renaming number array

#get list of ordered numbers
numbers_ordered = list(range(1, 176))
print(numbers_ordered)
#turn n umbers into strings
numbers_str = [str(num) for num in numbers_ordered]
#add 0 before single digits
for i in range(9):
    numbers_str[i] = "00" + numbers_str[i]
for i in range(9,99):
    numbers_str[i] = "0" + numbers_str[i]
#Shuffle numbers
random.shuffle(numbers_str)
print(numbers_str)
number_counter = 0

### Policy number variable that will be added to the html file
policynum = "var policynumber = '55648cefdd59e8c399e52c004f0bff6950f2fdf1a28ef8a2841766cf';"


### Excel file that contains the d values
excel = "Dnumbers.xlsx"
wb_obj = openpyxl.load_workbook(excel)
#Sheet 
sheetDnumbers = wb_obj["Dnumbers"]
row_Dnumber = 1

### Excel file for final lists
excel2 = "List.xlsx"
wb_obj2 = openpyxl.load_workbook(excel2)
#Sheet 
sheet_v1 = wb_obj2["v1"]
sheet_v2 = wb_obj2["v2"]
sheet_v3 = wb_obj2["v3"]
sheet_v4 = wb_obj2["v4"]
sheet_v5 = wb_obj2["v5"]
sheet_all = wb_obj2["all"]

sheets = {
     "v1" : 'wb_obj2["v1"]',
     "v2" : 'wb_obj2["v2"]',
     "v3" : 'wb_obj2["v3"]',
     "v4" : 'wb_obj2["v4"]',
     "v5" : 'wb_obj2["v5"]',
     "all" : 'wb_obj2["all"]'
     }

row_List = 2

html_counter = 1




###### v1 ######

#2d

src = os.getcwd()


for fname in os.listdir(src): 
    if (fname == "v1" or fname == "v2" or fname == "v3" or fname == "v4" or fname == "v5"):
        
        current_sheet = globals()["sheet_" + fname]

        # build the path to the v folders
        v_path = os.path.join(src, fname)
        print(v_path)
        
        #List excel iterators to save d values
        col_count = 2
        row_count = 2
        
        if os.path.isdir(v_path):
            for fname2 in os.listdir(v_path):  #do this in every v folder
    
                # build the path to the d folders
                d_path = os.path.join(v_path, fname2)
                print(d_path)
                
                #find files in folder
                folder_path = d_path
                dirs = os.listdir(folder_path)

                #Create new folder
                newpath = folder_path + "/new" 
                if not os.path.exists(newpath):
                    os.makedirs(newpath)
                    
            
                #Access files one by one
                for file in dirs:
                    if file.endswith(".html"):
                        file_name = file
                        #Open file for reading
                        template2d = open(folder_path + "\\" + file, "r")
                        content = template2d.read()
                        
                        
                        #Name the new file
                        htmlName_temp = numbers_str[number_counter] + "_" + fname + "_" + fname2 + "_4Fre5hBTC.html"
                        htmlName = htmlName_temp[:-5]
                        #Save name in excel
                        current_sheet['A' + str(row_count)] = htmlName
                        
                        
                        #Add policy number to html
                        content_split = content.split("</script>")
                        content = content_split[0] + "\n" + policynum + "\n" + "</script>" + content_split[1]
                        
                        #Count instances of d values
                        instance_count = content.count("getRandomIntDec(300, 500)")
                        instance_count += content.count("getRandomIntDec(300,500)")
                        
                        #Loop through instances
                        for inst in range(instance_count):
                            d_temp = sheetDnumbers['A' + str(row_Dnumber)].value
                            #Save d value in excel
                            current_sheet.cell(row=row_count, column=col_count).value = d_temp
                            col_count += 1
                            #wb_obj2.save("List.xlsx")
                            row_Dnumber += 1
                            #Replace placeholder with d value
                            content_temp_temp = content.replace("getRandomIntDec(300, 500)", d_temp, 1)
                            content_temp = content_temp_temp.replace("getRandomIntDec(300,500)", d_temp, 1)
                            content = content_temp
                            
                        #Save excel data
                        wb_obj2.save("List.xlsx")
                        col_count = 2 #reset column
                        row_count += 1 #
                            
                        #Create new html file and place it in the new folder 
                        path = fname + "/" + fname2 + "/new/" + htmlName + ".html"
                        newHtml = open(path, "w")
                        newHtml.write(content)
                        newHtml.close()
                        
                        keep_track = str(html_counter) + ": " + numbers_str[number_counter]
                        print(keep_track)
                        html_counter += 1
                        number_counter += 1
                        
                        
                        
        



       
            
            
        
        
        
        
        
        
        
        







#3d
#5d





###### v2 ######

#2d
#3d
#5d

###### v3 ######

#2d
#3d
#5d

###### v4 ######
#2d
#3d
#5d

###### v5 ######
#2d
#3d
#5d

